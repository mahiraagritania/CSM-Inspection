import streamlit as st
import os
import fitz  # PyMuPDF
from ultralytics import YOLO
import cv2
import numpy as np
import pytesseract
import pandas as pd
import camelot as cam
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment


# Direktori dasar untuk menyimpan file
base_upload_directory = r"C:\Users\madan\OneDrive - Institut Teknologi Bandung\Desktop\inspection_data\input"
croppedview_directory = r"C:\Users\madan\OneDrive - Institut Teknologi Bandung\Desktop\inspection_data\cropped_view"
croppeddim_directory = r"C:\Users\madan\OneDrive - Institut Teknologi Bandung\Desktop\inspection_data\cropped_dimension"
base_output_directory = r'C:\Users\madan\OneDrive - Institut Teknologi Bandung\Desktop\inspection_data\output'

# Path model YOLOv8n
model_view_path = r"C:\Users\madan\OneDrive - Institut Teknologi Bandung\Desktop\inspection_app\model\yolov8_view.pt"
model_annotation_path = r"C:\Users\madan\OneDrive - Institut Teknologi Bandung\Desktop\inspection_app\model\yolov8_annotation.pt"

# Membuat direktori dasar jika belum ada
if not os.path.exists(base_upload_directory):
    os.makedirs(base_upload_directory)

if not os.path.exists(croppedview_directory):
    os.makedirs(croppedview_directory)

if not os.path.exists(croppeddim_directory):
    os.makedirs(croppeddim_directory)

st.title("Inspection Sheet Generation")

# Dropdown untuk memilih nomor order
order_number = st.selectbox("Pilih nomor order", os.listdir(base_upload_directory))

if order_number:
    upload_directory = os.path.join(base_upload_directory, order_number)

    if not os.path.exists(upload_directory):
        st.error(f"Direktori {upload_directory} tidak ada.")
    else:
        st.header(f"File dalam direktori: {order_number}")
        files = [f for f in os.listdir(upload_directory) if f.lower().endswith('.pdf')]
        selected_file = st.selectbox("Pilih file PDF", files)

        if selected_file:
            file_path = os.path.join(upload_directory, selected_file)

            # Tombol "Mulai" untuk memulai konversi
            if st.button("Mulai"):
                # Fungsi untuk konversi halaman pertama PDF menjadi gambar
                def pdf_first_page_to_image(pdf_path, output_folder):
                    # Buka PDF
                    pdf_document = fitz.open(pdf_path)

                    # Ambil halaman pertama
                    first_page = pdf_document.load_page(0)
                    # Konversi halaman ke pixmap
                    pixmap = first_page.get_pixmap()

                    # Nama dasar PDF tanpa ekstensi
                    pdf_base_name = os.path.splitext(os.path.basename(pdf_path))[0]

                    # Simpan pixmap sebagai gambar dengan nama yang sama dengan file PDF
                    image_name = f"{pdf_base_name}.jpg"
                    image_path = os.path.join(output_folder, image_name)
                    pixmap.save(image_path)
                    
                    return image_path
        

                # Folder output untuk menyimpan gambar
                output_folder = upload_directory  # Menyimpan di direktori yang sama

                # Buat folder output jika belum ada
                if not os.path.exists(output_folder):
                    os.makedirs(output_folder)

                # Konversi halaman pertama PDF menjadi gambar
                image_path = pdf_first_page_to_image(file_path, output_folder)

                # Load pertama model YOLOv8n (untuk view)
                model_view = YOLO(model_view_path)

                # Run inference on the image
                results_view = model_view(image_path)

                # Load gambar asli
                img = cv2.imread(image_path)

                # Create direktori untuk simpan gambar yang di-crop oleh model view
                order_cropped_view_directory = os.path.join(croppedview_directory, order_number)
                if not os.path.exists(order_cropped_view_directory):
                    os.makedirs(order_cropped_view_directory)

                # Ekstrak bounding boxes dari hasil view
                boxes_view = results_view[0].boxes.xyxy.tolist()
                labels_view = results_view[0].boxes.cls.tolist()  # Dapatkan nama kelas

                class_names_view = ['view', 'infoblock']  # Nama kelas sesuai dengan indeks kelas

                # Iterasi melalui bounding boxes
                for i, box_view in enumerate(boxes_view):
                    x1_view, y1_view, x2_view, y2_view = box_view
                    # Crop objek menggunakan koordinat bounding box
                    cropped_object_view = img[int(y1_view):int(y2_view), int(x1_view):int(x2_view)]
                    # Dapatkan nama kelas yang sesuai dari daftar class_names
                    class_index_view = int(labels_view[i])
                    class_name_view = class_names_view[class_index_view]
                    # Simpan objek yang di-crop sebagai gambar dalam direktori cropped_view
                    cv2.imwrite(os.path.join(order_cropped_view_directory, f"{selected_file[:-4]}_{class_name_view}_{i}.jpg"), cropped_object_view)

                # Load kedua model YOLOv8n (untuk annotation)
                model_annotation = YOLO(model_annotation_path)

                # Create direktori untuk simpan gambar yang di-crop dan di-annotasi oleh model annotation
                order_cropped_dim_directory = os.path.join(croppeddim_directory, order_number)
                if not os.path.exists(order_cropped_dim_directory):
                    os.makedirs(order_cropped_dim_directory)

                # Daftar semua file JPG di direktori croppedview_directory
                jpg_files = [f for f in os.listdir(order_cropped_view_directory) if f.endswith('.jpg')]

                # Iterasi melalui file JPG
                for jpg_file in jpg_files:
                    # Konstruksi path lengkap ke file JPG input
                    input_image_path = os.path.join(order_cropped_view_directory, jpg_file)

                    # Jalankan inferensi pada gambar input
                    results_annotation = model_annotation(input_image_path)

                    # Muat gambar asli
                    img_annotation = cv2.imread(input_image_path)

                    # Ekstrak bounding boxes
                    boxes_annotation = results_annotation[0].boxes.xyxy.tolist()
                    labels_annotation = results_annotation[0].boxes.cls.tolist()  # Dapatkan nama kelas

                    class_names_annotation = ['datum', 'dimensi']  # Nama kelas sesuai dengan indeks kelas

                    # Iterasi melalui bounding boxes
                    for j, box_annotation in enumerate(boxes_annotation):
                        x1_annotation, y1_annotation, x2_annotation, y2_annotation = box_annotation
                        # Crop objek menggunakan koordinat bounding box
                        cropped_object_annotation = img_annotation[int(y1_annotation):int(y2_annotation), int(x1_annotation):int(x2_annotation)]
                        # Dapatkan nama kelas yang sesuai dari daftar class_names
                        class_index_annotation = int(labels_annotation[j])
                        class_name_annotation = class_names_annotation[class_index_annotation]
                        # Simpan objek yang di-crop sebagai gambar dalam direktori cropped_dimension
                        cv2.imwrite(os.path.join(order_cropped_dim_directory, f"{jpg_file[:-4]}_{class_name_annotation}_{j}.jpg"), cropped_object_annotation)

                # Lanjutan proses untuk pengolahan teks hasil OCR
                def is_green_font(image):
                    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
                    lower_green = np.array([67, 91, 114])  # Sesuaikan nilai-nilai ini sesuai dengan kebutuhan Anda
                    upper_green = np.array([87, 191, 214]) # Sesuaikan nilai-nilai ini sesuai dengan kebutuhan Anda
                    mask = cv2.inRange(hsv, lower_green, upper_green)
                    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    if contours:
                        max_contour_area = max(contours, key=cv2.contourArea)
                        contour_area = cv2.contourArea(max_contour_area)
                        green_threshold = 100
                        if contour_area > green_threshold:
                            return True
                    return False

                # Directory paths
                input_dir = order_cropped_dim_directory

                # List to store the results
                results = []

                # List all files in the input directory
                files = os.listdir(input_dir)
                for file in files:
                    # Construct full path to the input image file
                    input_image_path = os.path.join(input_dir, file)

                    # Load the image
                    image = cv2.imread(input_image_path)

                    # Resize the input image
                    resized_image = cv2.resize(image, None, fx=10, fy=10, interpolation=cv2.INTER_CUBIC)
                    image = resized_image

                    # Check image orientation
                    if image.shape[1] > image.shape[0]:
                        if is_green_font(image):
                            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                            blurred = cv2.GaussianBlur(gray, (9, 9), 0)
                            thresh = cv2.threshold(blurred, 100, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
                            kernel = np.ones((10, 10), np.uint8)  # Sesuaikan ukuran kernel sesuai kebutuhan
                            dilated = cv2.dilate(thresh, kernel, iterations=1)
                            contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                            contours = sorted(contours, key=lambda x: cv2.boundingRect(x)[0])
                            ocr_results = []
                            for contour in contours:
                                x, y, w, h = cv2.boundingRect(contour)
                                roi = gray[y:y+h, x:x+w]
                                text = pytesseract.image_to_string(roi, config="-c tessedit_char_whitelist='0123456789-.,Ø°PCRMN()DH+-spk' --psm 6")
                                ocr_results.append(text)
                            ocr_results_clean = [text.strip().replace('(', '').replace(')', '').replace(' ', '') for text in ocr_results if text.strip()]
                            ocr_results_merged = ' '.join(ocr_results_clean).replace('\n', '')
                            ocr_results_final = [text for text in ocr_results_merged.split(' ') if text]
                            final_text = ' '.join(ocr_results_final)
                            final_text = final_text.replace(' ', '')
                            results.extend(final_text.split())  # Pisahkan teks akhir dan tambahkan ke daftar hasil
                        else:
                            background_scale_factor = 2
                            canvas = np.ones((image.shape[0] * background_scale_factor, image.shape[1] * background_scale_factor, 3), dtype="uint8") * 255
                            start_x = int((canvas.shape[1] - image.shape[1]) / 2)
                            start_y = int((canvas.shape[0] - image.shape[0]) / 2)
                            canvas[start_y:start_y+image.shape[0], start_x:start_x+image.shape[1]] = image
                            gray = 255 - cv2.cvtColor(canvas, cv2.COLOR_BGR2GRAY)
                            blurred = cv2.GaussianBlur(gray, (9, 9), 0)
                            thresh = cv2.threshold(blurred, 100, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
                            text2 = pytesseract.image_to_string(thresh, config="-c tessedit_char_whitelist='0123456789-.,Ø°PCRMN()DH+-' --psm 6")
                            text2 = text2.replace(' ', '')
                            results.extend(text2.split())  # Pisahkan teks dan tambahkan ke daftar hasil
                    else:  # Orientasi potret
                        rotated_image = cv2.rotate(image, cv2.ROTATE_90_CLOCKWISE)
                        if is_green_font(rotated_image):
                            gray = cv2.cvtColor(rotated_image, cv2.COLOR_BGR2GRAY)
                            blurred = cv2.GaussianBlur(gray, (9, 9), 0)
                            thresh = cv2.threshold(blurred, 100, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
                            kernel = np.ones((10, 10), np.uint8)  # Sesuaikan ukuran kernel sesuai kebutuhan
                            dilated = cv2.dilate(thresh, kernel, iterations=1)
                            contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                            contours = sorted(contours, key=lambda x: cv2.boundingRect(x)[0])
                            ocr_results = []
                            for contour in contours:
                                x, y, w, h = cv2.boundingRect(contour)
                                roi = gray[y:y+h, x:x+w]
                                text = pytesseract.image_to_string(roi, config="-c tessedit_char_whitelist='0123456789-.,Ø°PCRMN()DH+-spk' --psm 6")
                                ocr_results.append(text)
                            ocr_results_clean = [text.strip().replace('(', '').replace(')', '').replace(' ', '') for text in ocr_results if text.strip()]
                            ocr_results_merged = ' '.join(ocr_results_clean).replace('\n', '')
                            ocr_results_final = [text for text in ocr_results_merged.split(' ') if text]
                            final_text2 = ' '.join(ocr_results_final)
                            final_text2 = final_text2.replace(' ', '')
                            results.extend(final_text2.split())  # Pisahkan teks akhir dan tambahkan ke daftar hasil
                        else:
                            background_scale_factor = 2
                            canvas = np.ones((rotated_image.shape[0] * background_scale_factor, rotated_image.shape[1] * background_scale_factor, 3), dtype="uint8") * 255
                            start_x = int((canvas.shape[1] - rotated_image.shape[1]) / 2)
                            start_y = int((canvas.shape[0] - rotated_image.shape[0]) / 2)
                            canvas[start_y:start_y+rotated_image.shape[0], start_x:start_x+rotated_image.shape[1]] = rotated_image
                            gray = 255 - cv2.cvtColor(canvas, cv2.COLOR_BGR2GRAY)
                            blurred = cv2.GaussianBlur(gray, (9, 9), 0)
                            thresh = cv2.threshold(blurred, 100, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
                            text3 = pytesseract.image_to_string(thresh, config="-c tessedit_char_whitelist='0123456789-.,Ø°PCRMN()DH+-' --psm 6")
                            text3 = text3.replace(' ', '')
                            results.extend(text3.split())  # Pisahkan teks dan tambahkan ke daftar hasil

                punctuation_to_remove = '-.,Ø°()+0'

                def is_mixed_alpha_punct_without_digit(item):
                    has_alpha = any(char.isalpha() for char in item)
                    has_punct = any(char in punctuation_to_remove for char in item)
                    has_digit = any(char.isdigit() for char in item)
                    return has_alpha and has_punct and not has_digit
                
                def post_process(results):
                    processed_result = []
                    next_item_index = 0
                    while next_item_index < len(results):

                        item = results[next_item_index]
                        if is_mixed_alpha_punct_without_digit(item):
                            next_item_index += 1
                            continue

                        if len(item) > 1 and item[0].isdigit() and item[1].isalpha():
                            next_item_index += 1
                            continue

                        if len(item) == 1 and item.isalpha():
                            next_item_index += 1
                            continue

                        if len(item) == 2 and item.startswith('0'):
                            item = '0' + ',' + item[1]

                        if item.startswith('P'):
                            item = 'D' + item[1:]

                        if item.startswith('H'):
                            item = 'D' + item[1:]

                        if all(char in punctuation_to_remove for char in item):
                            next_item_index += 1
                            continue

                        if len(item) == 3 and item[0] in ['D', 'R', ' C'] and item[1] in [0]:
                            item = item[0] + item[1] + ',' + item[2]

                        if len(item) == 1 and item in punctuation_to_remove:
                            next_item_index += 1
                            continue

                        if len(item) >= 2 and all(char.isalpha() for char in item):
                            next_item_index += 1
                            continue

                        if len(item) == 3 and item[0] in ['D', 'R']:
                            if item[1] in ['0']:
                                item = item[0] + item[1] + ',' + item[2]
                            elif item[1] in [',']:
                                item = item[0] + '0' + ','+ item[2]

                        if item.startswith('-'):
                            item = item[1:]
                        if len(item) == 2 and item[-1] == '0' and item[0] == 'R':
                                next_item_index += 1
                                continue

                        if '-' in item:
                            try:
                                before_dash, after_dash = item.split('-')
                                if after_dash[0].isdigit() and after_dash[0] == '0':
                                    item = before_dash + '-' + 'D' + after_dash[1:]
                            except ValueError:
                                pass

                        if item.startswith('.'):
                            item = item[1:]

                        if item.startswith('°'):
                            item = item[1:]

                        if item.endswith('.'):
                            item = item[:-1]

                        if item.endswith(','):
                            item = item[:-1]

                        # Logic to handle numbers starting with 1 and ending with (x) or (xx)
                        if item.startswith('1') and item.endswith(')'):
                            left_index = item.rfind('(')
                            if left_index != -1:
                                before_bracket = item[:left_index]
                                after_bracket = item[left_index + 1:-1]
                                if after_bracket.isdigit():
                                    item = before_bracket[-1] + '(' + after_bracket + ')'

                        # Logic to handle numbers starting with "("
                        if item.startswith('('):
                            if item.endswith(')'):
                                item = item
                            else:
                                item += ')'
                        
                        if item.startswith('0') and len(item)== 5:
                            item = 'D' + item[1:]

                        processed_result.append(item)
                        next_item_index += 1

                    return processed_result

                # Memproses hasil OCR untuk mendapatkan teks akhir
                processed_results = post_process(results)
                processed_results = [item.replace(',', '.') for item in processed_results]

                def extract_table_from_pdf(pdf_path):
                    tables = cam.read_pdf(pdf_path, pages='1', flavor='lattice')
                    df = tables[0].df
                    df = df.replace('\n', '', regex=True)
                    return df

                # Ekstrak tabel dari PDF dan tampilkan di Streamlit
                df = extract_table_from_pdf(file_path)
                print(df)

                part_name_cell = None
                for column in df.columns:
                    for index, cell_value in enumerate(df[column]):
                        if isinstance(cell_value, str) and cell_value.startswith('Part Name'):
                            part_name_cell = df.iloc[index, df.columns.get_loc(column)]
                            break
                    if part_name_cell:
                        break

                if part_name_cell:
                    part_name = part_name_cell.split(":")[1].strip()
                    if len(part_name) < 2:
                        def cari_part_name(row_index, column_index, dataframe):
                            if row_index + 1 < len(dataframe):
                                if not pd.isna(dataframe.iloc[row_index + 1, column_index]):
                                    return dataframe.iloc[row_index + 1, column_index]
                                else:
                                    return cari_part_name(row_index + 1, column_index, dataframe)
                            else:
                                return None

                        for index, row in df.iterrows():
                            for column_index, value in enumerate(row):
                                if value == 'Part Name :':
                                    part_name = cari_part_name(index, column_index, df)
                                    print("Part Name:", part_name)
                    else:
                        print("Part Name:", part_name)
                else:
                    print("Tidak ditemukan informasi 'Part Name'.")

                # 2. Mengambil informasi setelah "Part No."
                part_no_cell = None
                for column in df.columns:
                    for index, cell_value in enumerate(df[column]):
                        if isinstance(cell_value, str) and cell_value.startswith('Part No'):
                            part_no_cell = df.iloc[index, df.columns.get_loc(column)]
                            break
                    if part_no_cell:
                        break

                if part_no_cell:
                    part_no = part_no_cell.split(".")[1].strip()
                    print("Part No:", part_no)
                else:
                    print("Tidak ditemukan informasi 'Part No.'.")

                # 3. Mengambil informasi pada sel terdekat yang terisi di atas baris 'Qty'
                def cari_qty(row_index, column_index, dataframe):
                    if row_index - 1 >= 0:
                        if not pd.isna(dataframe.iloc[row_index - 1, column_index]):
                            return dataframe.iloc[row_index - 1, column_index]
                        else:
                            return cari_qty(row_index - 1, column_index, dataframe)
                    else:
                        return None

                for index, row in df.iterrows():
                    for column_index, value in enumerate(row):
                        if value == 'Qty':
                            quantity = cari_qty(index, column_index, df)
                            print("Quantity:", quantity)

                # 4. Mengambil informasi pada sel terdekat yang terisi di kanan 'Drawn by'
                def cari_drawn(row_index, column_index, dataframe):
                    if column_index + 1 < len(dataframe.columns):
                        if not pd.isna(dataframe.iloc[row_index, column_index + 1]):
                            return dataframe.iloc[row_index, column_index + 1]
                        else:
                            return cari_drawn(row_index, column_index + 1, dataframe)
                    else:
                        return None

                for index, row in df.iterrows():
                    for column_index, value in enumerate(row):
                        if isinstance(value, str) and 'Drawn by' in value:
                            drawn_by = cari_drawn(index, column_index, df)
                            print("Drawn by:", drawn_by)

                # 5. Mencari informasi material
                def cari_material(row_index, column_index, dataframe):
                    # Jika baris di atasnya ada
                    if row_index - 1 >= 0:
                        # Cek apakah sel di atasnya terisi
                        if not pd.isna(dataframe.iloc[row_index - 1, column_index]):
                            # Jika terisi, kembalikan nilai dari sel di atasnya
                            return dataframe.iloc[row_index - 1, column_index]
                        # Jika tidak terisi, lanjutkan pencarian ke atas
                        elif not pd.isna(dataframe.iloc[row_index - 1, column_index - 1]):  # Cari ke kiri
                            return dataframe.iloc[row_index - 1, column_index - 1]
                        elif not pd.isna(dataframe.iloc[row_index - 2, column_index]):  # Cari ke atas satu lagi
                            return dataframe.iloc[row_index - 2, column_index]
                        else:
                            # Jika masih tidak ditemukan, lanjutkan pencarian ke kiri atau ke atas lagi
                            return cari_material(row_index - 1, column_index, dataframe)
                    else:
                        # Jika mencapai batas atas, kembalikan None
                        return None

                # Temukan indeks baris dan kolom untuk sel yang mengandung 'material''
                for index, row in df.iterrows():
                    for column_index, value in enumerate(row):
                        if value == 'Material':
                            # Periksa apakah sel di atasnya berisi informasi
                            jenis_material = cari_material(index, column_index, df)
                            if jenis_material is not None:
                                # Periksa jika jenis material terdiri dari lebih dari satu kata
                                if len(jenis_material.split()) > 1:
                                    print("Material:", jenis_material)
                                else:
                                    print("Material:", jenis_material)


                # 6. Mencari informasi tanggal
                def find_custom_date_format(cell_value):
                    pattern = r'\b\d{1,2}/\d{2}/\d{4}\b'
                    match = re.search(pattern, cell_value)
                    if match:
                        return match.group()
                    else:
                        return None

                for index, row in df.iterrows():
                    for column in df.columns:
                        cell_value = str(row[column])
                        custom_date = find_custom_date_format(cell_value)
                        if custom_date:
                            print("Tanggal:", custom_date)

                # 7. Mencari informasi dimension
                def cari_dimensi(row_index, column_index, dataframe):
                    if row_index - 1 >= 0:
                        if not pd.isna(dataframe.iloc[row_index - 1, column_index]):
                            return dataframe.iloc[row_index - 1, column_index]
                        else:
                            return cari_dimensi(row_index - 1, column_index, dataframe)
                    else:
                        return None

                for index, row in df.iterrows():
                    for column_index, value in enumerate(row):
                        if value == 'Dimension':
                            dimension = cari_dimensi(index, column_index, df)
                            print("Dimensi:", dimension)

                # Langkah 1: Mengambil/load excel bernama "Template"
                wb = load_workbook('C:\\Users\\madan\\OneDrive - Institut Teknologi Bandung\\Desktop\\inspection_app\\model\\Template.xlsx')
                sheet = wb.active  # Mengambil sheet aktif, bisa juga menggunakan wb['Sheet1'] jika nama sheet diketahui

                # Langkah 2: Memasukkan variabel "part_name" pada sel C5
                sheet['C5'] = part_name

                # Langkah 3: Memasukkan variabel "part_no" pada sel C4
                sheet['C4'] = part_no

                # Langkah 4: Memasukkan variabel "quantity" pada sel C6
                sheet['C6'] = quantity

                # Langkah 6: Memasukkan variabel "dimensi" pada sel mulai dari B10
                for i, value in enumerate(processed_results):
                    cell = 'B{}'.format(10 + i)  # Mulai dari baris 10
                    sheet[cell] = value

                # Langkah 7: Load Tolerance Guidance
                wb_toleransi = load_workbook('C:\\Users\\madan\\OneDrive - Institut Teknologi Bandung\\Desktop\\inspection_app\\model\\Tolerance Guidance.xlsx')
                tolerance_type = "Medium"

                # Langkah 8: Define tolerance type
                tolerance_rows = {
                    "Fine": 3,
                    "Medium": 4,
                    "Coarse": 5,
                }

                # Mendapatkan baris toleransi berdasarkan tipe
                tolerance_row = tolerance_rows[tolerance_type]

                # Langkah 9: Menuliskan nomor berurutan dari 1 pada sel A10 sesuai dengan sel C yang terisi
                last_row = len(processed_results) + 9  # Perhatikan penyesuaian baris terakhir
                for i in range(10, last_row + 1):  # Dimulai dari baris 10
                    sheet['A{}'.format(i)] = i - 9  # Nomor berurutan dimulai dari 1 pada baris 10

                # Langkah 10: Memasukkan nilai toleransi pada sel mulai dari C10
                for i in range(10, 10 + len(processed_results)):
                    cell_value = sheet['B{}'.format(i)].value
                    if cell_value:
                        if len(cell_value) > 6 and '+' in cell_value:
                            try:
                                dimensi_digit, tol = cell_value.split('+')
                                if 'M' in dimensi_digit:
                                    dimensi_digit = dimensi_digit.split('M')[0]
                                if 'H' in dimensi_digit:
                                    dimensi_digit = dimensi_digit.split('H')[0]
                                try:
                                    dimensi = float(dimensi_digit)
                                    tolerance_value = float(tol)
                                    sheet['C{}'.format(i)] = tolerance_value
                                    sheet['D{}'.format(i)] = dimensi + tolerance_value
                                    sheet['E{}'.format(i)] = dimensi - tolerance_value
                                except ValueError:
                                    sheet['C{}'.format(i)] = None
                                    sheet['D{}'.format(i)] = None
                                    sheet['E{}'.format(i)] = None
                            except ValueError:
                                sheet['C{}'.format(i)] = None
                                sheet['D{}'.format(i)] = None
                                sheet['E{}'.format(i)] = None

                        elif cell_value.startswith("(") and cell_value.endswith(")"):
                            dimensi = float(cell_value[1:-1])
                            sheet['B{}'.format(i)] = cell_value
                            if dimensi >= 0 and dimensi < 6:
                                column_index = 'C'
                            elif dimensi >= 6 and dimensi < 30:
                                column_index = 'D'
                            elif dimensi >= 30 and dimensi < 120:
                                column_index = 'E'
                            elif dimensi >= 120 and dimensi < 315:
                                column_index = 'F'
                            elif dimensi >= 315 and dimensi < 1000:
                                column_index = 'G'
                            elif dimensi >= 1000 and dimensi <= 2000:
                                column_index = 'H'
                            tolerance_value = wb_toleransi['Sheet1'][f'{column_index}{tolerance_row}'].value
                            sheet['C{}'.format(i)] = tolerance_value
                            sheet['D{}'.format(i)] = dimensi + tolerance_value
                            sheet['E{}'.format(i)] = dimensi - tolerance_value
                        
                        else:
                            if cell_value.replace('.', '').isdigit():
                                if len(cell_value) > 5:
                                    dimensi = float(cell_value.replace('.', ''))
                                else:
                                    dimensi = float(cell_value)                         
                            elif any(char in cell_value for char in 'PCRMDHspk'):
                                if '-' in cell_value:
                                    back_strip = cell_value.split('-')[-1]
                                    digit_value = ''.join(filter(str.isdigit, back_strip))
                                    try:
                                        dimensi = float(digit_value)
                                    except ValueError:
                                        dimensi = cell_value
                                        sheet['C{}'.format(i)] = None
                                        sheet['D{}'.format(i)] = None
                                        sheet['E{}'.format(i)] = None
                                else:
                                    digit_value = cell_value[1:]
                                    try:
                                        dimensi = float(digit_value)
                                    except ValueError:
                                        dimensi = cell_value
                            elif '°' in cell_value:
                                dimensi = float(''.join(filter(str.isdigit, cell_value[:-1])))
                            else:
                                dimensi = None

                            if dimensi is not None:
                                if isinstance(dimensi, str):
                                    sheet['C{}'.format(i)] = None
                                    sheet['D{}'.format(i)] = None
                                    sheet['E{}'.format(i)] = None
                                else:
                                    if dimensi >= 0 and dimensi < 6:
                                        column_index = 'C'
                                    elif dimensi >= 6 and dimensi < 30:
                                        column_index = 'D'
                                    elif dimensi >= 30 and dimensi < 120:
                                        column_index = 'E'
                                    elif dimensi >= 120 and dimensi < 315:
                                        column_index = 'F'
                                    elif dimensi >= 315 and dimensi < 1000:
                                        column_index = 'G'
                                    elif dimensi >= 1000 and dimensi <= 2000:
                                        column_index = 'H'
                                    elif dimensi > 2000:
                                        column_index = 'I'
                                    tolerance_value = wb_toleransi['Sheet1'][f'{column_index}{tolerance_row}'].value
                                    sheet['C{}'.format(i)] = tolerance_value
                                    sheet['D{}'.format(i)] = dimensi + tolerance_value
                                    sheet['E{}'.format(i)] = dimensi - tolerance_value

                # Langkah 11: Membuat seluruh sel dari A10 sampai E44 jadi rata tengah (center)
                for row in sheet.iter_rows(min_row=10, max_row=44, min_col=1, max_col=5):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # Langkah 12: Membuat sel C5, C6, dan E6 jadi rata kiri (left)
                sheet['C5'].alignment = Alignment(horizontal='left', vertical='center')
                sheet['C6'].alignment = Alignment(horizontal='left', vertical='center')
                sheet['E6'].alignment = Alignment(horizontal='left', vertical='center')

                # Path untuk menyimpan file
                order_folder_path = os.path.join(base_output_directory, str(order_number))
                if not os.path.exists(order_folder_path):
                    os.makedirs(order_folder_path)

                # Path lengkap untuk menyimpan file dengan nama yang sesuai dengan nama pdf
                file_name = os.path.splitext(os.path.basename(file_path))[0]
                output_path = os.path.join(order_folder_path, f'{file_name}.xlsx')

                # Menyimpan file dengan nama yang sesuai dengan variabel order_number
                wb.save(output_path)
                st.success("Proses telah selesai.")
                st.success(f"Tolong melakukan pengecekan lembar inspeksi di {output_path}")